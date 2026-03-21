"""
scraper_agricole_italie.py
Scraping d'entreprises agricoles italiennes sur PagineGialle.it
Utilise Playwright (vrai navigateur Chromium) pour contourner AWS WAF.
"""

# ── Encodage UTF-8 console Windows ───────────────────────────────────────────
import sys, subprocess
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

# ── Auto-install dépendances ──────────────────────────────────────────────────
def _pip(*pkgs):
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", *pkgs])

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except ImportError:
    print("Installation de playwright…")
    _pip("playwright")
    subprocess.check_call([sys.executable, "-m", "playwright", "install", "chromium"])
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installation de openpyxl…"); _pip("openpyxl")
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

try:
    from tqdm import tqdm
except ImportError:
    print("Installation de tqdm…"); _pip("tqdm")
    from tqdm import tqdm

try:
    from bs4 import BeautifulSoup
except ImportError:
    print("Installation de beautifulsoup4…"); _pip("beautifulsoup4")
    from bs4 import BeautifulSoup

# ── Imports standard ─────────────────────────────────────────────────────────
import csv
import random
import re
import time
from dataclasses import dataclass, asdict
from typing import Optional

# ─────────────────────────────────────────────────────────────────────────────
# Configuration
# ─────────────────────────────────────────────────────────────────────────────

KEYWORDS = [
    "azienda agricola",
    "cooperativa agricola",
    "masseria",
    "frantoio",
    "cantina",
]

REGIONS = [
    "Puglia",
    "Sicilia",
    "Emilia-Romagna",
    "Veneto",
    "Calabria",
    "Campania",
    "Toscana",
    "Lazio",
]

MAX_PAGES_PER_SEARCH = 5
DELAY_MIN = 2.5
DELAY_MAX = 4.5

OUTPUT_EXCEL = "resultats_agricoles_italie.xlsx"
OUTPUT_CSV   = "resultats_agricoles_italie.csv"

# ─────────────────────────────────────────────────────────────────────────────
# Modèle de données
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class Entreprise:
    nom:       str = ""
    ville:     str = ""
    province:  str = ""
    region:    str = ""
    telephone: str = ""
    email:     str = ""
    site_web:  str = ""
    source_url:str = ""

    def cle_dedup(self) -> str:
        n = re.sub(r"\W+", "", self.nom.lower())
        t = re.sub(r"\D", "", self.telephone)
        return f"{n}|{t}"


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _pause():
    time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))


def _slug(text: str) -> str:
    return text.strip().lower().replace(" ", "-")


def _build_url(keyword: str, region: str, page: int = 1) -> str:
    kw  = _slug(keyword)
    reg = _slug(region)
    base = f"https://www.paginegialle.it/ricerca/{kw}/{reg}"
    return base if page == 1 else f"{base}?pg={page}"


def _extract_email(text: str) -> str:
    m = re.search(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", text)
    return m.group(0).lower() if m else ""


# ─────────────────────────────────────────────────────────────────────────────
# Parsing HTML (BeautifulSoup sur contenu déjà chargé par Playwright)
# ─────────────────────────────────────────────────────────────────────────────

def _parse_block(block, region: str) -> Optional[Entreprise]:
    e = Entreprise(region=region)

    # URL profil
    link = block.select_one("a.remove_blank_for_app[href]")
    if link:
        e.source_url = link.get("href", "").strip()

    # Nom
    nom_tag = block.select_one("h2.search-itm__rag")
    if nom_tag:
        for span in nom_tag.select("span.icon-check"):
            span.decompose()
        e.nom = nom_tag.get_text(" ", strip=True)
    if not e.nom:
        return None

    # Adresse → ville + province
    adr_tag = block.select_one("div.search-itm__adr")
    if adr_tag:
        adr_text = adr_tag.get_text(" ", strip=True)
        m = re.search(r"\b(\d{5})\s+([^(,\n]+?)\s*\(([A-Z]{2})\)", adr_text)
        if m:
            e.ville    = m.group(2).strip()
            e.province = m.group(3).strip()
        else:
            m2 = re.search(r"([^,\-\d]{3,}?)\s*\(([A-Z]{2})\)", adr_text)
            if m2:
                e.ville    = m2.group(1).strip()
                e.province = m2.group(2).strip()

    # Téléphone (div masqué révélé par Playwright)
    phone_div = block.select_one("div.search-itm__phone")
    if phone_div:
        phone_text = phone_div.get_text(" ", strip=True)
        m = re.search(r"[\+\d][\d\s\-\.]{6,}", phone_text)
        if m:
            e.telephone = re.sub(r"[\s\-\.]", "", m.group(0))

    # Email inline
    for a in block.find_all("a", href=True):
        if a["href"].startswith("mailto:"):
            e.email = a["href"][7:].split("?")[0].lower()
            break
    if not e.email:
        e.email = _extract_email(block.get_text(" "))

    # Site web (lien externe)
    for a in block.find_all("a", href=True):
        href = a["href"]
        if (href.startswith("http")
                and "paginegialle" not in href
                and "italiaonline" not in href
                and "plug.it" not in href
                and "google" not in href
                and "facebook" not in href):
            e.site_web = href.rstrip("/")
            break

    return e


def _parse_page(html: str, region: str) -> list[Entreprise]:
    soup = BeautifulSoup(html, "html.parser")
    blocks = soup.select("div.search-itm")
    results = []
    for block in blocks:
        e = _parse_block(block, region)
        if e:
            results.append(e)
    return results


# ─────────────────────────────────────────────────────────────────────────────
# Playwright : navigateur headless
# ─────────────────────────────────────────────────────────────────────────────

def _wait_for_results(page, timeout_ms: int = 12000) -> bool:
    """Attend que les résultats soient chargés (div.search-itm ou message 'nessun risultato')."""
    try:
        page.wait_for_selector(
            "div.search-itm, .listing-res__numresults",
            timeout=timeout_ms,
        )
        return True
    except PWTimeout:
        return False


def scrape_keyword_region(
    pw_page,
    keyword: str,
    region: str,
) -> list[Entreprise]:
    results = []
    print(f"\n  🔍 '{keyword}' dans {region}")

    for page_num in range(1, MAX_PAGES_PER_SEARCH + 1):
        url = _build_url(keyword, region, page_num)

        try:
            pw_page.goto(url, wait_until="domcontentloaded", timeout=30000)
        except Exception as ex:
            print(f"    → Erreur navigation page {page_num} : {ex}")
            break

        loaded = _wait_for_results(pw_page)
        if not loaded:
            print(f"    → Page {page_num} : timeout ou vide, arrêt.")
            break

        html = pw_page.content()
        batch = _parse_page(html, region)

        if not batch:
            print(f"    → Page {page_num} : 0 résultat parsé, arrêt.")
            break

        results.extend(batch)
        print(f"    → Page {page_num} : {len(batch)} entreprise(s) ({len(results)} cumulé)")

        # Vérifier page suivante
        soup = BeautifulSoup(html, "html.parser")
        has_next = bool(
            soup.select_one("a[rel='next']")
            or soup.select_one("[data-tr='listing-pagination-next']")
            or soup.find("a", string=re.compile(r"successiv|Avanti", re.I))
        )
        if not has_next:
            break

        _pause()

    return results


# ─────────────────────────────────────────────────────────────────────────────
# Enrichissement emails via /contatti
# ─────────────────────────────────────────────────────────────────────────────

def enrich_contatti(pw_page, entreprises: list[Entreprise]) -> None:
    sans = [e for e in entreprises if not e.email and e.site_web]
    if not sans:
        return

    print(f"\n📧 Enrichissement /contatti ({len(sans)} sites à visiter)…")
    for e in tqdm(sans, desc="Contatti", unit="site"):
        base = e.site_web.rstrip("/")
        for path in ("/contatti", "/contact", "/chi-siamo"):
            try:
                pw_page.goto(
                    f"{base}{path}",
                    wait_until="domcontentloaded",
                    timeout=6000,   # 6s max par page
                )
                html = pw_page.content()
                soup = BeautifulSoup(html, "html.parser")

                for a in soup.find_all("a", href=True):
                    if a["href"].startswith("mailto:"):
                        e.email = a["href"][7:].split("?")[0].lower()
                        break

                if not e.email:
                    e.email = _extract_email(soup.get_text(" "))

                if e.email:
                    break
            except Exception:
                pass
            time.sleep(0.5)  # pause courte entre pages du même site


# ─────────────────────────────────────────────────────────────────────────────
# Dédoublonnage
# ─────────────────────────────────────────────────────────────────────────────

def deduplicate(entreprises: list[Entreprise]) -> list[Entreprise]:
    seen: set[str] = set()
    unique = []
    for e in entreprises:
        k = e.cle_dedup()
        if k and k not in seen:
            seen.add(k)
            unique.append(e)
    return unique


# ─────────────────────────────────────────────────────────────────────────────
# Export Excel
# ─────────────────────────────────────────────────────────────────────────────

COLONNES = ["Nom", "Ville", "Province", "Région", "Téléphone", "Email", "Site web", "Source URL"]
CHAMPS   = ["nom", "ville", "province", "region", "telephone", "email", "site_web", "source_url"]


def export_excel(entreprises: list[Entreprise], filepath: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Entreprises agricoles IT"

    h_fill  = PatternFill("solid", fgColor="1F4E79")
    h_font  = Font(bold=True, color="FFFFFF", size=11)
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    for ci, col_name in enumerate(COLONNES, 1):
        c = ws.cell(row=1, column=ci, value=col_name)
        c.fill = h_fill; c.font = h_font; c.alignment = h_align

    fills = [
        PatternFill("solid", fgColor="DEEAF1"),
        PatternFill("solid", fgColor="FFFFFF"),
    ]
    for ri, e in enumerate(entreprises, 2):
        fill = fills[ri % 2]
        for ci, champ in enumerate(CHAMPS, 1):
            cell = ws.cell(row=ri, column=ci, value=getattr(e, champ, ""))
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 4, 60)
    ws.freeze_panes = "A2"
    wb.save(filepath)
    print(f"✅ Excel exporté → {filepath}  ({len(entreprises)} lignes)")


def export_csv(entreprises: list[Entreprise], filepath: str) -> None:
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=CHAMPS, extrasaction="ignore")
        w.writeheader()
        for e in entreprises:
            w.writerow(asdict(e))
    print(f"✅ CSV exporté   → {filepath}  ({len(entreprises)} lignes)")


# ─────────────────────────────────────────────────────────────────────────────
# Point d'entrée principal
# ─────────────────────────────────────────────────────────────────────────────

def run(regions: list[str] = REGIONS, test_mode: bool = False) -> None:
    all_entreprises: list[Entreprise] = []
    total_combos = len(KEYWORDS) * len(regions)

    print(f"\n{'='*60}")
    print(f"  SCRAPER AGRICOLE ITALIE — PagineGialle.it")
    print(f"  {len(KEYWORDS)} mots-clés × {len(regions)} région(s) = {total_combos} recherches")
    if test_mode:
        print("  MODE TEST : Puglia uniquement")
    print(f"{'='*60}\n")

    combos = [(kw, reg) for reg in regions for kw in KEYWORDS]

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/122.0.0.0 Safari/537.36"
            ),
            locale="it-IT",
            viewport={"width": 1280, "height": 900},
        )
        page = context.new_page()

        # Pré-charger la page d'accueil pour résoudre le cookie WAF
        print("Initialisation du navigateur (résolution AWS WAF)…")
        try:
            page.goto("https://www.paginegialle.it/", wait_until="networkidle", timeout=20000)
            _pause()
        except Exception:
            pass

        combo_bar = tqdm(combos, desc="Recherches", unit="combo")
        for keyword, region in combo_bar:
            combo_bar.set_postfix({"kw": keyword[:18], "reg": region})
            batch = scrape_keyword_region(page, keyword, region)
            all_entreprises.extend(batch)
            _pause()

        print(f"\n📊 Total brut : {len(all_entreprises)} entreprise(s)")
        all_entreprises = deduplicate(all_entreprises)
        print(f"📊 Après dédoublonnage : {len(all_entreprises)} entreprise(s) uniques")

        # Enrichissement emails /contatti
        enrich_contatti(page, all_entreprises)

        context.close()
        browser.close()

    print()
    export_excel(all_entreprises, OUTPUT_EXCEL)
    export_csv(all_entreprises, OUTPUT_CSV)

    avec_email = sum(1 for e in all_entreprises if e.email)
    avec_site  = sum(1 for e in all_entreprises if e.site_web)
    avec_tel   = sum(1 for e in all_entreprises if e.telephone)
    total      = max(len(all_entreprises), 1)

    print(f"\n{'='*60}")
    print(f"  RÉSUMÉ FINAL")
    print(f"  Total entreprises uniques : {len(all_entreprises)}")
    print(f"  Avec téléphone : {avec_tel}  ({avec_tel*100//total} %)")
    print(f"  Avec email     : {avec_email}  ({avec_email*100//total} %)")
    print(f"  Avec site web  : {avec_site}  ({avec_site*100//total} %)")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Scraper entreprises agricoles IT (Playwright)")
    parser.add_argument("--test",    action="store_true", help="Mode test : Puglia uniquement")
    parser.add_argument("--regions", nargs="+", default=None, help="Régions spécifiques")
    args = parser.parse_args()

    if args.test:
        run(regions=["Puglia"], test_mode=True)
    elif args.regions:
        run(regions=args.regions)
    else:
        run()

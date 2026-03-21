"""
Génère un fichier Excel avec 200 entreprises italiennes dans les secteurs
qui utilisent le plus le Decreto Flussi pour recruter des travailleurs étrangers.
"""
import sys
import os

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ─── DONNÉES : 200 entreprises italiennes ────────────────────────────────────
# Colonnes : Ragione Sociale, Settore, Tipo, Regione, Provincia, Sito Web, Note

AZIENDE = [
    # ── AGRICOLTURA & ALIMENTARE (50) ─────────────────────────────────────────
    ("Apofruit Italia Coop.", "Agricoltura / Alimentare", "Cooperativa agricola", "Emilia-Romagna", "FC", "www.apofruit.it", "Una delle maggiori cooperative frutticole d'Italia"),
    ("Conserve Italia Coop.", "Agricoltura / Alimentare", "Cooperativa agroalimentare", "Emilia-Romagna", "BO", "www.conserveitalia.it", "Leader nella trasformazione di frutta e verdura"),
    ("Agrintesa", "Agricoltura / Alimentare", "Cooperativa agricola", "Emilia-Romagna", "RA", "www.agrintesa.it", "Cooperativa frutticola del Ravennate"),
    ("Orogel SpA", "Agricoltura / Alimentare", "Industria alimentare", "Emilia-Romagna", "FC", "www.orogel.it", "Leader nel surgelato vegetale in Italia"),
    ("Bonifiche Ferraresi SpA", "Agricoltura / Alimentare", "Azienda agricola quotata", "Emilia-Romagna", "FE", "www.bonificeferraresi.it", "Più grande azienda agricola quotata in borsa in Italia"),
    ("Amadori Group", "Agricoltura / Alimentare", "Industria avicola", "Emilia-Romagna", "FC", "www.amadori.it", "Leader italiano nella filiera avicola"),
    ("Granarolo SpA", "Agricoltura / Alimentare", "Industria lattiero-casearia", "Emilia-Romagna", "BO", "www.granarolo.it", "Primo gruppo lattiero-caseario italiano"),
    ("Cremonini SpA", "Agricoltura / Alimentare", "Industria carni", "Emilia-Romagna", "MO", "www.cremonini.com", "Primo produttore di carni bovine in Europa"),
    ("Fileni Alimentare", "Agricoltura / Alimentare", "Industria avicola", "Marche", "AN", "www.fileni.it", "Secondo produttore avicolo in Italia"),
    ("AIA SpA", "Agricoltura / Alimentare", "Industria avicola", "Veneto", "VR", "www.aia.it", "Azienda Italiana Allevamenti, leader settore"),
    ("Rana SpA", "Agricoltura / Alimentare", "Pasta fresca industriale", "Veneto", "VR", "www.giovannirana.it", "Pasta fresca leader in Italia"),
    ("Zuegg SpA", "Agricoltura / Alimentare", "Industria frutta", "Trentino-Alto Adige", "TN", "www.zuegg.it", "Succhi di frutta e confetture"),
    ("Cavit", "Agricoltura / Alimentare", "Cooperativa vinicola", "Trentino-Alto Adige", "TN", "www.cavit.it", "Prima cantina cooperativa del Trentino"),
    ("Mezzacorona", "Agricoltura / Alimentare", "Cooperativa vinicola", "Trentino-Alto Adige", "TN", "www.mezzacorona.it", "Grande cooperativa vinicola trentina"),
    ("Masi Agricola SpA", "Agricoltura / Alimentare", "Azienda vinicola", "Veneto", "VR", "www.masi.it", "Storica cantina della Valpolicella"),
    ("Cantina di Soave", "Agricoltura / Alimentare", "Cooperativa vinicola", "Veneto", "VR", "www.cantinadisoave.it", "Più grande cantina cooperativa del Veneto"),
    ("Antinori SpA", "Agricoltura / Alimentare", "Azienda vinicola", "Toscana", "FI", "www.antinori.it", "Famiglia vitivinicola dal 1385"),
    ("Frescobaldi SpA", "Agricoltura / Alimentare", "Azienda vinicola", "Toscana", "FI", "www.frescobaldi.it", "Storica cantina toscana"),
    ("Ruffino SpA", "Agricoltura / Alimentare", "Azienda vinicola", "Toscana", "FI", "www.ruffino.it", "Vini Chianti e Brunello"),
    ("Castello Banfi", "Agricoltura / Alimentare", "Azienda vinicola", "Toscana", "SI", "www.castellobanfi.com", "Celebre cantina di Montalcino"),
    ("Fontanafredda SpA", "Agricoltura / Alimentare", "Azienda vinicola", "Piemonte", "CN", "www.fontanafredda.it", "Storica cantina di Serralunga d'Alba"),
    ("Gaja", "Agricoltura / Alimentare", "Azienda vinicola", "Piemonte", "CN", "www.gaja.com", "Produttore di Barolo e Barbaresco"),
    ("Planeta SpA", "Agricoltura / Alimentare", "Azienda vinicola", "Sicilia", "AG", "www.planeta.it", "Vini siciliani di eccellenza"),
    ("Tasca d'Almerita", "Agricoltura / Alimentare", "Azienda vinicola", "Sicilia", "PA", "www.tascadalmerita.it", "Cantina siciliana di riferimento"),
    ("Donnafugata SpA", "Agricoltura / Alimentare", "Azienda vinicola", "Sicilia", "TP", "www.donnafugata.it", "Vini di Pantelleria e Sicilia"),
    ("Cantine Settesoli", "Agricoltura / Alimentare", "Cooperativa vinicola", "Sicilia", "AG", "www.settesoli.it", "Grande cooperativa vinicola siciliana"),
    ("Oranfrizer", "Agricoltura / Alimentare", "Lavorazione agrumi", "Sicilia", "SR", "www.oranfrizer.it", "Arance rosse di Sicilia"),
    ("Feudi di San Gregorio", "Agricoltura / Alimentare", "Azienda vinicola", "Campania", "AV", "www.feudi.it", "Vini Irpinia, eccellenza campana"),
    ("Cantine Due Palme", "Agricoltura / Alimentare", "Cooperativa vinicola", "Puglia", "BR", "www.cantineduepalme.it", "Cooperativa pugliese con migliaia di soci"),
    ("Torrevento SpA", "Agricoltura / Alimentare", "Azienda vinicola", "Puglia", "BA", "www.torrevento.it", "Vini Castel del Monte, Puglia"),
    ("Rivera SpA", "Agricoltura / Alimentare", "Azienda vinicola", "Puglia", "BT", "www.rivera.it", "Storica cantina pugliese"),
    ("Masseria Il Frantoio", "Agricoltura / Alimentare", "Masseria / agriturismo", "Puglia", "BR", "www.masseriailfrantoio.it", "Masseria con produzione olivicola"),
    ("De Cecco SpA", "Agricoltura / Alimentare", "Pastificio industriale", "Abruzzo", "CH", "www.dececco.it", "Pastificio artigianale industriale"),
    ("Mutti SpA", "Agricoltura / Alimentare", "Industria conserve pomodoro", "Emilia-Romagna", "PR", "www.mutti-parma.com", "Leader italiano nel pomodoro"),
    ("Barilla Group", "Agricoltura / Alimentare", "Industria alimentare", "Emilia-Romagna", "PR", "www.barilla.com", "Primo gruppo alimentare italiano"),
    ("Monini SpA", "Agricoltura / Alimentare", "Oleificio / olio extravergine", "Umbria", "PG", "www.monini.com", "Leader italiano nell'olio d'oliva"),
    ("Carapelli Firenze SpA", "Agricoltura / Alimentare", "Oleificio", "Toscana", "FI", "www.carapelli.it", "Storico marchio oleario toscano"),
    ("Luigi Lavazza SpA", "Agricoltura / Alimentare", "Industria caffè", "Piemonte", "TO", "www.lavazza.it", "Secondo torrefattore al mondo"),
    ("Illycaffè SpA", "Agricoltura / Alimentare", "Industria caffè", "Friuli-Venezia Giulia", "TS", "www.illy.com", "Caffè di alta qualità"),
    ("Ferrero Group", "Agricoltura / Alimentare", "Industria dolciaria", "Piemonte", "CN", "www.ferrero.com", "Nutella, Kinder: secondo in Italia"),
    ("Surgital SpA", "Agricoltura / Alimentare", "Pasta surgelata", "Emilia-Romagna", "RA", "www.surgital.it", "Leader pasta fresca surgelata"),
    ("Produttori del Barbaresco", "Agricoltura / Alimentare", "Cooperativa vinicola", "Piemonte", "CN", "www.produttorididelbarbaresco.com", "Cooperativa DOCG Barbaresco"),
    ("Fratelli Carli SpA", "Agricoltura / Alimentare", "Oleificio", "Liguria", "IM", "www.fratelli-carli.it", "Olio Carli, storico marchio ligure"),
    ("Princes Italia", "Agricoltura / Alimentare", "Conserve pomodoro", "Campania", "SA", "www.princes.it", "Concentrato e polpa di pomodoro"),
    ("Mastroberardino SpA", "Agricoltura / Alimentare", "Azienda vinicola", "Campania", "AV", "www.mastroberardino.com", "Fiano di Avellino, Taurasi"),
    ("Oleificio Zucchi", "Agricoltura / Alimentare", "Oleificio", "Lombardia", "LO", "www.oleificiozucchi.it", "Olio e condimenti, storica azienda"),
    ("Consorzio Produttori Vini Manduria", "Agricoltura / Alimentare", "Cooperativa vinicola", "Puglia", "TA", "www.cpvmanduria.it", "Primitivo di Manduria DOC"),
    ("Cantina Valpolicella Negrar", "Agricoltura / Alimentare", "Cooperativa vinicola", "Veneto", "VR", "www.cantinenegrar.it", "Amarone e Ripasso della Valpolicella"),
    ("Vivai Cooperativi Rauscedo", "Agricoltura / Alimentare", "Vivaio viticolo", "Friuli-Venezia Giulia", "PN", "www.vivairauscedo.com", "Maggior produttore mondiale barbatelle"),
    ("Bertani Domains", "Agricoltura / Alimentare", "Azienda vinicola", "Veneto", "VR", "www.bertani.net", "Amarone storico, Veneto"),

    # ── EDILIZIA & COSTRUZIONI (30) ───────────────────────────────────────────
    ("Webuild S.p.A.", "Edilizia / Costruzioni", "Grande impresa costruzioni", "Lombardia", "MI", "www.webuildgroup.com", "Più grande costruttore italiano, ex Salini-Impregilo"),
    ("Impresa Pizzarotti & C. SpA", "Edilizia / Costruzioni", "Grande impresa costruzioni", "Emilia-Romagna", "PR", "www.pizzarotti.it", "Secondo costruttore italiano"),
    ("Itinera SpA", "Edilizia / Costruzioni", "Costruzioni infrastrutture", "Piemonte", "AL", "www.itinera.com", "Gruppo Gavio, infrastrutture stradali"),
    ("Rizzani de Eccher SpA", "Edilizia / Costruzioni", "Grande impresa costruzioni", "Friuli-Venezia Giulia", "UD", "www.rizzanideeccher.com", "Terzo costruttore italiano"),
    ("CMC di Ravenna", "Edilizia / Costruzioni", "Cooperativa costruzioni", "Emilia-Romagna", "RA", "www.cmcravenna.com", "Grande cooperativa di costruzione"),
    ("Grandi Lavori Fincosit SpA", "Edilizia / Costruzioni", "Grande impresa costruzioni", "Lazio", "RM", "www.grandilavori.com", "Tunnel e infrastrutture sotterranee"),
    ("Ghella SpA", "Edilizia / Costruzioni", "Costruzioni gallerie", "Lazio", "RM", "www.ghella.com", "Leader mondiale nelle gallerie TBM"),
    ("CCC Cooperativa Costruzioni", "Edilizia / Costruzioni", "Cooperativa costruzioni", "Emilia-Romagna", "BO", "www.ccc.coop", "Cooperativa di costruzione Bologna"),
    ("Toto Costruzioni Generali", "Edilizia / Costruzioni", "Grande impresa costruzioni", "Abruzzo", "CH", "www.totocostruzioni.it", "Autostrade e infrastrutture"),
    ("Salcef Group SpA", "Edilizia / Costruzioni", "Costruzioni ferroviarie", "Lazio", "RM", "www.salcef.it", "Leader manutenzione ferroviaria"),
    ("Techbau SpA", "Edilizia / Costruzioni", "Costruzioni civili", "Lombardia", "MI", "www.techbau.it", "Costruzioni civili e industriali"),
    ("Cossi Costruzioni SpA", "Edilizia / Costruzioni", "Costruzioni civili", "Lombardia", "BG", "www.cossicosstruzioni.it", "Edilizia civile e industriale"),
    ("ICM SpA", "Edilizia / Costruzioni", "Grande impresa costruzioni", "Veneto", "VI", "www.icm.it", "Impresa Costruzioni Maltauro"),
    ("Mantovani SpA", "Edilizia / Costruzioni", "Costruzioni idrauliche", "Veneto", "VE", "www.mantovani.it", "Grandi opere idrauliche"),
    ("Coopsette", "Edilizia / Costruzioni", "Cooperativa costruzioni", "Emilia-Romagna", "RE", "www.coopsette.coop", "Cooperativa edile reggiana"),
    ("Guerrato SpA", "Edilizia / Costruzioni", "Costruzioni industriali", "Veneto", "RO", "www.guerrato.it", "Impianti industriali e civili"),
    ("ICOP SpA", "Edilizia / Costruzioni", "Geotecnica", "Friuli-Venezia Giulia", "UD", "www.icopfoundations.com", "Fondazioni speciali e geotecnica"),
    ("Cimolai SpA", "Edilizia / Costruzioni", "Costruzioni metalliche", "Friuli-Venezia Giulia", "PN", "www.cimolai.com", "Strutture metalliche e ponti"),
    ("Sicim SpA", "Edilizia / Costruzioni", "Costruzioni pipeline", "Emilia-Romagna", "PR", "www.sicim.eu", "Pipeline oil & gas mondiale"),
    ("Bonatti SpA", "Edilizia / Costruzioni", "Costruzioni oil & gas", "Emilia-Romagna", "PR", "www.bonatti.it", "Infrastrutture energetiche"),
    ("Saipem SpA", "Edilizia / Costruzioni", "Ingegneria offshore", "Lombardia", "MI", "www.saipem.com", "Grande gruppo E&C energia"),
    ("Maire Tecnimont SpA", "Edilizia / Costruzioni", "Ingegneria industriale", "Lombardia", "MI", "www.mairetecnimont.com", "Impianti petrolchimici"),
    ("Gemmo SpA", "Edilizia / Costruzioni", "Impianti tecnologici", "Veneto", "VI", "www.gemmo.it", "Facility management e impianti"),
    ("Cooperativa Muratori di Ravenna", "Edilizia / Costruzioni", "Cooperativa costruzioni", "Emilia-Romagna", "RA", "www.cmcravenna.com", "Cooperativa storica emiliana"),
    ("Società Italiana Condotte d'Acqua", "Edilizia / Costruzioni", "Costruzioni idrauliche", "Lazio", "RM", "www.condotte.it", "Acquedotti e opere idrauliche"),
    ("Lodigiani SpA", "Edilizia / Costruzioni", "Grande impresa costruzioni", "Lombardia", "MI", "www.lodigiani.it", "Costruzioni civili e infrastrutture"),
    ("Cambielli Edilfriuli SpA", "Edilizia / Costruzioni", "Distribuzione edile", "Friuli-Venezia Giulia", "PN", "www.cambielli.it", "Materiali edili, distribuzione"),
    ("Edilco SpA", "Edilizia / Costruzioni", "Costruzioni civili", "Lazio", "RM", "www.edilco.it", "Edilizia residenziale e commerciale"),
    ("Ferrante Costruzioni", "Edilizia / Costruzioni", "Costruzioni civili", "Campania", "NA", "—", "Costruzioni civili Campania"),
    ("Costruzioni Generali Gilardi", "Edilizia / Costruzioni", "Costruzioni civili", "Piemonte", "TO", "—", "Edilizia civile e industriale Piemonte"),

    # ── TURISMO & ALBERGHIERO (35) ────────────────────────────────────────────
    ("Starhotels SpA", "Turismo / Alberghiero", "Catena alberghiera", "Toscana", "FI", "www.starhotels.com", "29 hotel 4-5 stelle in Italia e estero"),
    ("Bettoja Hotels", "Turismo / Alberghiero", "Catena alberghiera", "Lazio", "RM", "www.bettojahotels.it", "Hotel storici a Roma dal 1875"),
    ("Baglioni Hotels", "Turismo / Alberghiero", "Catena alberghiera lusso", "Lombardia", "MI", "www.baglionihotels.com", "9 hotel extra-lusso Italia e estero"),
    ("TH Resorts", "Turismo / Alberghiero", "Catena resort", "Lazio", "RM", "www.thinncollection.com", "Villaggi e resort in Italia"),
    ("Alpitour World", "Turismo / Alberghiero", "Gruppo turistico integrato", "Piemonte", "TO", "www.alpitourworld.it", "Primo operatore turistico italiano"),
    ("Valtur SpA", "Turismo / Alberghiero", "Villaggi vacanze", "Lazio", "RM", "www.valtur.it", "Catena di villaggi vacanze"),
    ("SINA Hotels", "Turismo / Alberghiero", "Catena alberghiera", "Toscana", "FI", "www.sinahotels.com", "Hotel di lusso in Italia"),
    ("Una Hotels & Resorts", "Turismo / Alberghiero", "Catena alberghiera", "Toscana", "FI", "www.unahotels.it", "45 hotel in tutta Italia"),
    ("Blu Hotels", "Turismo / Alberghiero", "Catena alberghiera", "Lombardia", "BS", "www.bluhotels.it", "Hotel e resort in Italia"),
    ("Domina Hotels & Resorts", "Turismo / Alberghiero", "Catena alberghiera", "Sicilia", "PA", "www.dominahotels.com", "Hotel in Sicilia e nel mondo"),
    ("Mangia's Resorts", "Turismo / Alberghiero", "Catena resort", "Sicilia", "PA", "www.mangias.com", "Resort esclusivi in Sicilia"),
    ("Aeroviaggi SpA", "Turismo / Alberghiero", "Catena resort", "Sicilia", "PA", "www.aeroviaggi.it", "Villaggi vacanze in Sicilia"),
    ("Boscolo Hotels", "Turismo / Alberghiero", "Catena alberghiera", "Veneto", "PD", "www.boscolohotels.com", "Hotel 5 stelle in Italia"),
    ("Terme di Saturnia SpA", "Turismo / Alberghiero", "Terme e resort", "Toscana", "GR", "www.termedisaturnia.it", "Resort termale di lusso"),
    ("Terme Euganee SpA", "Turismo / Alberghiero", "Terme e resort", "Veneto", "PD", "www.termeuganee.it", "Complesso termale Abano Terme"),
    ("Costa Crociere SpA", "Turismo / Alberghiero", "Crociere", "Liguria", "GE", "www.costacrociere.it", "Prima compagnia crocieristica italiana"),
    ("MSC Crociere", "Turismo / Alberghiero", "Crociere", "Campania", "NA", "www.msccrociere.it", "Seconda compagnia crocieristica mondiale"),
    ("Grimaldi Lines", "Turismo / Alberghiero", "Traghetti e crociere", "Campania", "NA", "www.grimaldi-lines.com", "Ferry e crociere Mediterraneo"),
    ("Club Méditerranée Italia", "Turismo / Alberghiero", "Villaggi vacanze", "Lombardia", "MI", "www.clubmed.it", "Villaggi all-inclusive in Italia"),
    ("Vivosa Apulia Resort", "Turismo / Alberghiero", "Resort pugliese", "Puglia", "LE", "www.vivosa.it", "Resort esclusivo Salento"),
    ("Grand Hotel Timeo", "Turismo / Alberghiero", "Hotel lusso", "Sicilia", "ME", "www.fourseasons.com", "Four Seasons a Taormina"),
    ("Grand Hotel Tremezzo", "Turismo / Alberghiero", "Hotel lusso lago", "Lombardia", "CO", "www.grandhoteltremezzo.com", "Hotel 5 stelle Lago di Como"),
    ("Hotel Cipriani Venezia", "Turismo / Alberghiero", "Hotel lusso", "Veneto", "VE", "www.belmond.com", "Belmond Hotel Cipriani"),
    ("Rocco Forte Hotels Italy", "Turismo / Alberghiero", "Catena hotel lusso", "Lazio", "RM", "www.roccofortehotels.com", "Hotel di lusso in Italia"),
    ("NH Hotels Italia", "Turismo / Alberghiero", "Catena alberghiera", "Lombardia", "MI", "www.nh-hotels.it", "Catena Minor Hotels in Italia"),
    ("Marriott Hotels Italia", "Turismo / Alberghiero", "Catena alberghiera", "Lombardia", "MI", "www.marriott.it", "Marriott, Westin, Sheraton in Italia"),
    ("Hilton Hotels Italia", "Turismo / Alberghiero", "Catena alberghiera", "Lazio", "RM", "www.hilton.com", "Hilton, DoubleTree, Hampton in Italia"),
    ("Accor Hotels Italia", "Turismo / Alberghiero", "Catena alberghiera", "Lombardia", "MI", "www.accor.com", "Novotel, Ibis, Sofitel in Italia"),
    ("Masseria Torre Coccaro", "Turismo / Alberghiero", "Masseria resort", "Puglia", "BR", "www.torrecoccaro.com", "Masseria di lusso Fasano"),
    ("IHG Hotels Italia", "Turismo / Alberghiero", "Catena alberghiera", "Lombardia", "MI", "www.ihg.com", "Holiday Inn, Crowne Plaza in Italia"),
    ("Radisson Hotels Italia", "Turismo / Alberghiero", "Catena alberghiera", "Lazio", "RM", "www.radissonhotels.com", "Radisson Blu e RED in Italia"),
    ("Best Western Hotels Italia", "Turismo / Alberghiero", "Catena alberghiera", "Lombardia", "MI", "www.bestwestern.it", "Catena con 170+ hotel in Italia"),
    ("Federalberghi — Garda", "Turismo / Alberghiero", "Consorzio alberghiero", "Veneto", "VR", "www.gardaincoming.it", "Consorzio hotel Lago di Garda"),
    ("Acaya Golf Resort & Spa", "Turismo / Alberghiero", "Resort golf", "Puglia", "LE", "www.acayagolfresort.it", "Resort con golf in Salento"),
    ("Grand Hotel Villa Serbelloni", "Turismo / Alberghiero", "Hotel lusso lago", "Lombardia", "CO", "www.villaserbelloni.com", "Hotel 5 stelle Bellagio"),

    # ── RISTORAZIONE & CATERING (20) ──────────────────────────────────────────
    ("Autogrill SpA", "Ristorazione / Catering", "Ristorazione autostradale", "Lombardia", "MI", "www.autogrill.com", "Leader europeo ristorazione viaggiatori"),
    ("Chef Express SpA", "Ristorazione / Catering", "Ristorazione stazioni e aerei", "Emilia-Romagna", "MO", "www.chefexpress.it", "Gruppo Cremonini, stazioni e aeroporti"),
    ("Roadhouse SpA", "Ristorazione / Catering", "Catena ristoranti", "Emilia-Romagna", "MO", "www.roadhouse.it", "Gruppo Cremonini, steakhouse"),
    ("La Piadineria SpA", "Ristorazione / Catering", "Catena fast food", "Lombardia", "MI", "www.lapiadineria.com", "Catena con 350+ punti vendita"),
    ("Old Wild West (Cigierre)", "Ristorazione / Catering", "Catena ristoranti", "Lombardia", "MI", "www.cigierre.it", "Tex-mex e burger, 100+ locali"),
    ("Marr SpA", "Ristorazione / Catering", "Distribuzione ristorazione", "Emilia-Romagna", "RN", "www.marr.it", "Leader distribuzione foodservice"),
    ("Eataly SpA", "Ristorazione / Catering", "Ristorazione e retail", "Piemonte", "TO", "www.eataly.net", "Eccellenze alimentari italiane"),
    ("Compass Group Italia", "Ristorazione / Catering", "Catering industriale", "Lombardia", "MI", "www.compass-group.it", "Leader mondiale nel catering"),
    ("Sodexo Italia", "Ristorazione / Catering", "Catering e FM", "Lombardia", "MI", "www.sodexo.com/it", "Ristorazione collettiva e FM"),
    ("Elior Italia", "Ristorazione / Catering", "Ristorazione collettiva", "Lombardia", "MI", "www.elior.it", "Catering scuole, aziende, ospedali"),
    ("Vivenda SpA", "Ristorazione / Catering", "Catering ospedaliero", "Lazio", "RM", "www.vivenda.it", "Catering strutture sanitarie"),
    ("Areas Italia (MyChef)", "Ristorazione / Catering", "Ristorazione autostradale", "Lombardia", "MI", "www.areas.eu", "Punti ristoro autostrade e aeroporti"),
    ("Camst Soc.Coop.", "Ristorazione / Catering", "Cooperativa ristorazione", "Emilia-Romagna", "BO", "www.camst.it", "Catering scolastico e aziendale"),
    ("Cir Food Coop", "Ristorazione / Catering", "Cooperativa ristorazione", "Emilia-Romagna", "RE", "www.cirfood.com", "Ristorazione collettiva cooperativa"),
    ("Rossopomodoro SpA", "Ristorazione / Catering", "Catena pizzerie", "Campania", "NA", "www.rossopomodoro.it", "Catena di pizzerie napoletane"),
    ("Doppio Malto SpA", "Ristorazione / Catering", "Catena ristoranti", "Lombardia", "CO", "www.doppiomalto.com", "Birrerie e ristoranti"),
    ("Tre Spade Group", "Ristorazione / Catering", "Catering industriale", "Lombardia", "MI", "—", "Ristorazione aziendale"),
    ("McDonald's Italia", "Ristorazione / Catering", "Fast food", "Lombardia", "MI", "www.mcdonalds.it", "650+ ristoranti in Italia"),
    ("Burger King Italia", "Ristorazione / Catering", "Fast food", "Lombardia", "MI", "www.burgerking.it", "200+ ristoranti in Italia"),
    ("Restalia Italia", "Ristorazione / Catering", "Catena ristorazione", "Lombardia", "MI", "www.restalia.es", "TGB e Pizza Hut in Italia"),

    # ── LOGISTICA & TRASPORTI (25) ────────────────────────────────────────────
    ("DHL Express Italia", "Logistica / Trasporti", "Corriere espresso", "Lombardia", "MI", "www.dhl.com/it", "Leader mondiale nella logistica"),
    ("BRT Corriere Espresso SpA", "Logistica / Trasporti", "Corriere espresso", "Emilia-Romagna", "BO", "www.brt.it", "Ex Bartolini, secondo in Italia"),
    ("GLS Italy SpA", "Logistica / Trasporti", "Corriere espresso", "Lombardia", "LO", "www.gls-italy.com", "Terzo corriere in Italia"),
    ("TNT Italia SpA (FedEx)", "Logistica / Trasporti", "Corriere espresso", "Lombardia", "MI", "www.tnt.com/it", "FedEx-TNT, spedizioni internazionali"),
    ("SDA Express Courier SpA", "Logistica / Trasporti", "Corriere espresso", "Lazio", "RM", "www.sda.it", "Poste Italiane, 4° in Italia"),
    ("Arcese Trasporti SpA", "Logistica / Trasporti", "Trasporto internazionale", "Veneto", "PD", "www.arcese.com", "Logistica internazionale"),
    ("Fercam SpA", "Logistica / Trasporti", "Trasporto internazionale", "Trentino-Alto Adige", "BZ", "www.fercam.com", "Spedizioni internazionali"),
    ("Savino Del Bene SpA", "Logistica / Trasporti", "Spedizioni internazionali", "Toscana", "FI", "www.savinodelbene.com", "Freight forwarding globale"),
    ("Italmondo SpA", "Logistica / Trasporti", "Trasporti e logistica", "Veneto", "PD", "www.italmondo.it", "Logistica integrata"),
    ("Geodis Italia SpA", "Logistica / Trasporti", "Logistica integrata", "Lombardia", "MI", "www.geodis.com/it", "Gruppo SNCF, logistica globale"),
    ("DB Schenker Italia", "Logistica / Trasporti", "Logistica integrata", "Lombardia", "MI", "www.dbschenker.com/it", "Logistica Deutsche Bahn"),
    ("Kuehne+Nagel Italia", "Logistica / Trasporti", "Spedizioni internazionali", "Lombardia", "MI", "www.kuehne-nagel.com/it", "Freight forwarding globale"),
    ("CEVA Logistics Italia", "Logistica / Trasporti", "Logistica integrata", "Lombardia", "MI", "www.cevalogistics.com", "Contract logistics e freight"),
    ("Rhenus Logistics Italia", "Logistica / Trasporti", "Logistica integrata", "Veneto", "VR", "www.rhenus.com/it", "Logistica multimodale"),
    ("Fiege Italia SpA", "Logistica / Trasporti", "Contract logistics", "Piemonte", "AL", "www.fiege.com/it", "Logistica fashion e retail"),
    ("DSV Italia SpA", "Logistica / Trasporti", "Spedizioni internazionali", "Lombardia", "MI", "www.dsv.com/it", "Terzo spedizioniere mondiale"),
    ("XPO Logistics Italia", "Logistica / Trasporti", "Trasporto su strada", "Lombardia", "MI", "www.xpo.com/it", "Trasporto e logistica"),
    ("Palletways Italia", "Logistica / Trasporti", "Distribuzione pallet", "Lombardia", "MI", "www.palletways.com/it", "Rete distribuzione pallet"),
    ("Italtrans SpA", "Logistica / Trasporti", "Trasporti conto terzi", "Lombardia", "BG", "www.italtrans.it", "Autotrasporto e logistica"),
    ("Dachser Italia", "Logistica / Trasporti", "Logistica europea", "Lombardia", "MI", "www.dachser.com/it", "Logistica merci europee"),
    ("Poste Italiane SpA", "Logistica / Trasporti", "Postal & logistica", "Lazio", "RM", "www.poste.it", "Primo operatore postale italiano"),
    ("Trenitalia SpA", "Logistica / Trasporti", "Trasporto ferroviario", "Lazio", "RM", "www.trenitalia.com", "Ferrovie dello Stato, passeggeri"),
    ("ITA Airways SpA", "Logistica / Trasporti", "Trasporto aereo", "Lazio", "RM", "www.itaairways.com", "Compagnia aerea nazionale italiana"),
    ("TNT Post Italia", "Logistica / Trasporti", "Servizi postali", "Lombardia", "MI", "www.fedex.com/it", "Servizi postali e corriere"),
    ("Luís Simões Italia", "Logistica / Trasporti", "Trasporti temperature controllata", "Lombardia", "MI", "www.luis-simoes.it", "Logistica refrigerata"),

    # ── MECCANICA & INDUSTRIA (20) ────────────────────────────────────────────
    ("Stellantis Italy (ex FCA)", "Meccanica / Industria", "Automotive", "Piemonte", "TO", "www.stellantis.com", "Fiat, Alfa Romeo, Lancia, Jeep"),
    ("Pirelli & C. SpA", "Meccanica / Industria", "Pneumatici", "Lombardia", "MI", "www.pirelli.com", "Leader mondiale pneumatici premium"),
    ("Brembo SpA", "Meccanica / Industria", "Sistemi frenanti", "Lombardia", "BG", "www.brembo.com", "Leader mondiale freni auto e moto"),
    ("CNH Industrial Italia", "Meccanica / Industria", "Macchine agricole e veicoli", "Piemonte", "TO", "www.cnhindustrial.com", "Case IH, New Holland, Iveco"),
    ("Comau SpA", "Meccanica / Industria", "Automazione industriale", "Piemonte", "TO", "www.comau.com", "Robot e automazione industriale"),
    ("Interpump Group SpA", "Meccanica / Industria", "Pompe idrauliche", "Emilia-Romagna", "RE", "www.interpumpgroup.it", "Leader mondiale pompe ad alta pressione"),
    ("Elica SpA", "Meccanica / Industria", "Cappe cucina", "Marche", "AN", "www.elica.com", "Leader mondiale cappe aspiranti"),
    ("Ariston Thermo Group", "Meccanica / Industria", "Caldaie e scaldabagni", "Marche", "AN", "www.aristonthermo.com", "Leader mondiale riscaldamento acqua"),
    ("Smeg SpA", "Meccanica / Industria", "Elettrodomestici", "Emilia-Romagna", "RE", "www.smeg.it", "Design e elettrodomestici premium"),
    ("De'Longhi Group SpA", "Meccanica / Industria", "Elettrodomestici", "Veneto", "TV", "www.delonghi.com", "Macchine caffè e elettrodomestici"),
    ("Gewiss SpA", "Meccanica / Industria", "Materiale elettrico", "Lombardia", "BG", "www.gewiss.com", "Impianti elettrici e domotica"),
    ("Sacmi Imola SC", "Meccanica / Industria", "Macchine industriali", "Emilia-Romagna", "BO", "www.sacmi.com", "Macchine per ceramica e bevande"),
    ("IMA Group SpA", "Meccanica / Industria", "Macchine confezionatrici", "Emilia-Romagna", "BO", "www.ima.it", "Macchine per farmaceutica e tè"),
    ("SCM Group SpA", "Meccanica / Industria", "Macchine legno e compositi", "Emilia-Romagna", "RN", "www.scmgroup.com", "Leader macchine per il legno"),
    ("Datalogic SpA", "Meccanica / Industria", "Tecnologie di lettura", "Emilia-Romagna", "BO", "www.datalogic.com", "Barcode scanner e automazione"),
    ("Marposs SpA", "Meccanica / Industria", "Metrologia industriale", "Emilia-Romagna", "BO", "www.marposs.com", "Sistemi di misura per industria"),
    ("Ferrari N.V.", "Meccanica / Industria", "Automotive lusso", "Emilia-Romagna", "MO", "www.ferrari.com", "Supercar di lusso"),
    ("Lamborghini SpA", "Meccanica / Industria", "Automotive lusso", "Emilia-Romagna", "BO", "www.lamborghini.com", "Supercar sportive"),
    ("Marchesini Group SpA", "Meccanica / Industria", "Macchine confezionatrici", "Emilia-Romagna", "BO", "www.marchesini.com", "Packaging farmaceutico"),
    ("Cefla SpA", "Meccanica / Industria", "Macchine verniciatura", "Emilia-Romagna", "BO", "www.cefla.com", "Macchine per legno e finitura"),

    # ── AGENZIE LAVORO & STAFFING (20) ───────────────────────────────────────
    ("Adecco Italia SpA", "Agenzie Lavoro / Staffing", "Agenzia per il lavoro", "Lombardia", "MI", "www.adecco.it", "Prima agenzia lavoro in Italia"),
    ("Manpower Italia SpA", "Agenzie Lavoro / Staffing", "Agenzia per il lavoro", "Lombardia", "MI", "www.manpower.it", "Seconda agenzia lavoro in Italia"),
    ("Randstad Italia SpA", "Agenzie Lavoro / Staffing", "Agenzia per il lavoro", "Lombardia", "MI", "www.randstad.it", "Terza agenzia lavoro in Italia"),
    ("Gi Group Holding", "Agenzie Lavoro / Staffing", "Agenzia per il lavoro", "Lombardia", "MI", "www.gigroup.it", "Primo gruppo HR italiano"),
    ("Synergie Italia", "Agenzie Lavoro / Staffing", "Agenzia per il lavoro", "Emilia-Romagna", "BO", "www.synergie-italia.it", "Agenzia specializzata in somministrazione"),
    ("Etjca SpA", "Agenzie Lavoro / Staffing", "Agenzia per il lavoro", "Lombardia", "MI", "www.etjca.it", "Agenzia per il lavoro storica"),
    ("Umana SpA", "Agenzie Lavoro / Staffing", "Agenzia per il lavoro", "Veneto", "VE", "www.umana.it", "Somministrazione e ricerca"),
    ("Ali SpA", "Agenzie Lavoro / Staffing", "Agenzia per il lavoro", "Veneto", "PD", "www.alispa.it", "Agenzia lavoro nordest"),
    ("Openjobmetis SpA", "Agenzie Lavoro / Staffing", "Agenzia per il lavoro", "Lombardia", "VA", "www.openjobmetis.it", "APL quotata in Borsa"),
    ("Orienta SpA", "Agenzie Lavoro / Staffing", "Agenzia per il lavoro", "Lazio", "RM", "www.orienta.net", "Somministrazione e ricerca"),
    ("Lavorint SpA", "Agenzie Lavoro / Staffing", "Agenzia per il lavoro", "Lazio", "RM", "www.lavorint.it", "APL specializzata decreto flussi"),
    ("Kelly Services Italia", "Agenzie Lavoro / Staffing", "Agenzia per il lavoro", "Lombardia", "MI", "www.kellyservices.it", "Profili tecnici e amministrativi"),
    ("Generazione Vincente SpA", "Agenzie Lavoro / Staffing", "Agenzia per il lavoro", "Lazio", "RM", "www.generazionevincente.it", "APL con focus su lavoratori stranieri"),
    ("Temporary SpA", "Agenzie Lavoro / Staffing", "Agenzia per il lavoro", "Lombardia", "MI", "www.temporary.it", "Somministrazione lavoro"),
    ("In.Te.S.A. SpA", "Agenzie Lavoro / Staffing", "Agenzia per il lavoro", "Lombardia", "MI", "www.intesa.mi.it", "APL industria e logistica"),
    ("Iama Risorse Umane", "Agenzie Lavoro / Staffing", "Agenzia per il lavoro", "Veneto", "PD", "www.iama.it", "APL Nordest"),
    ("Quanta SpA", "Agenzie Lavoro / Staffing", "Agenzia per il lavoro", "Lombardia", "MI", "www.quanta.it", "Somministrazione e outplacement"),
    ("Obiettivo Lavoro SpA", "Agenzie Lavoro / Staffing", "Agenzia per il lavoro", "Lombardia", "MI", "www.obiettivo-lavoro.it", "APL con 300+ filiali in Italia"),
    ("Trenkwalder Italia", "Agenzie Lavoro / Staffing", "Agenzia per il lavoro", "Veneto", "VR", "www.trenkwalder.com/it", "APL centro-europea con presenza in Italia"),
    ("Groupe Actual Italia", "Agenzie Lavoro / Staffing", "Agenzia per il lavoro", "Lombardia", "MI", "www.groupeactual.eu", "APL franco-italiana"),
]

print(f"Totale aziende: {len(AZIENDE)}")

# ─── CRÉATION EXCEL ──────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Aziende Decreto Flussi"

# Couleurs
HDR_FILL   = PatternFill("solid", fgColor="1F5C99")   # bleu foncé
ALT_FILL   = PatternFill("solid", fgColor="EAF2FB")   # bleu très clair
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
BORDER_COLOR = "C0C0C0"

thin = Side(style="thin", color=BORDER_COLOR)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = [
    "N°", "Ragione Sociale", "Settore", "Tipo Impresa",
    "Regione", "Provincia", "Sito Web", "Note / Fonte"
]

# Largeurs colonnes
col_widths = [5, 35, 28, 30, 22, 12, 32, 48]

for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Titre principal
ws.merge_cells("A1:H1")
title_cell = ws["A1"]
title_cell.value = "200 AZIENDE ITALIANE — PRINCIPALI SETTORI DECRETO FLUSSI"
title_cell.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
title_cell.fill = PatternFill("solid", fgColor="0D3B6E")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# Sous-titre
ws.merge_cells("A2:H2")
sub = ws["A2"]
sub.value = (
    "Aziende operanti nei settori che più utilizzano il Decreto Flussi per il reclutamento di lavoratori stranieri. "
    "Fonte: ricerca web 2024-2025 (Confagricoltura, ANCE, Federalberghi, FIPE, ecc.)"
)
sub.font = Font(italic=True, size=9, color="555555")
sub.fill = PatternFill("solid", fgColor="D6E8F7")
sub.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 22

# En-têtes
ws.append([])  # row 3 = blank spacer — on va utiliser row 3 pour headers
# Rewrite row 3
for col_i, h in enumerate(HEADERS, 1):
    c = ws.cell(row=3, column=col_i, value=h)
    c.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
ws.row_dimensions[3].height = 22

# Données
settori_vus = {}
row_num = 4
for idx, az in enumerate(AZIENDE, 1):
    settore = az[1]
    fill = ALT_FILL if (idx % 2 == 0) else WHITE_FILL

    row_data = [idx] + list(az)
    for col_i, val in enumerate(row_data, 1):
        c = ws.cell(row=row_num, column=col_i, value=val)
        c.fill = fill
        c.border = border
        c.alignment = Alignment(vertical="center", wrap_text=(col_i == 8))
        if col_i == 1:
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(size=9, bold=True)
        elif col_i == 2:
            c.font = Font(size=9, bold=True)
        else:
            c.font = Font(size=9)
    ws.row_dimensions[row_num].height = 15
    row_num += 1

# Freeze header
ws.freeze_panes = "B4"

# Filtre auto
ws.auto_filter.ref = f"A3:H{row_num - 1}"

# ── Feuille de stats par secteur ─────────────────────────────────────────────
ws2 = wb.create_sheet("Statistiche per Settore")
ws2.column_dimensions["A"].width = 32
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 50

# Titre
ws2.merge_cells("A1:C1")
t = ws2["A1"]
t.value = "DISTRIBUZIONE PER SETTORE"
t.font = Font(bold=True, size=13, color="FFFFFF")
t.fill = PatternFill("solid", fgColor="0D3B6E")
t.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 26

# Headers stats
for ci, h in enumerate(["Settore", "N° Aziende", "Principali profili ricercati"], 1):
    c = ws2.cell(row=2, column=ci, value=h)
    c.font = Font(bold=True, color="FFFFFF", size=10)
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border
ws2.row_dimensions[2].height = 20

profili = {
    "Agricoltura / Alimentare": "Braccianti agricoli, operai raccolta, potatori, operatori di cantina, addetti lavorazione alimentare",
    "Edilizia / Costruzioni": "Muratori, carpentieri, ferraioli, gruisti, operatori macchine movimento terra, idraulici, elettricisti",
    "Turismo / Alberghiero": "Camerieri ai piani, receptionist, portabagagli, addetti SPA, animatori, bagnini, barman",
    "Ristorazione / Catering": "Aiuto-cuoco, cameriere di sala, pizzaiolo, lavapiatti, addetti mensa, runner",
    "Logistica / Trasporti": "Magazzinieri, carrellisti, autisti patente CE/C, operatori picking, addetti smistamento",
    "Meccanica / Industria": "Operatori CNC, saldatori, assemblatori, addetti catena di montaggio, manutentori",
    "Agenzie Lavoro / Staffing": "Intermediatari: gestiscono richieste di nulla osta per conto di clienti in tutti i settori",
}

# Compte par secteur
from collections import Counter
cnt = Counter(az[1] for az in AZIENDE)

r = 3
for i, (sett, n) in enumerate(sorted(cnt.items()), 1):
    fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
    for ci, val in enumerate([sett, n, profili.get(sett, "—")], 1):
        c = ws2.cell(row=r, column=ci, value=val)
        c.fill = fill
        c.border = border
        c.font = Font(size=9)
        c.alignment = Alignment(vertical="center", wrap_text=(ci == 3))
    ws2.row_dimensions[r].height = 30
    r += 1

# Total
for ci, val in enumerate(["TOTALE", sum(cnt.values()), ""], 1):
    c = ws2.cell(row=r, column=ci, value=val)
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F5C99")
    c.border = border
    c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[r].height = 18

# ── Sauvegarde ───────────────────────────────────────────────────────────────
out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                   "decreto_flussi_200_aziende.xlsx")
wb.save(out)
print(f"\nExcel sauvegardé : {out}")
print(f"Aziende totali   : {len(AZIENDE)}")
print(f"Settori          : {len(cnt)}")
for s, n in sorted(cnt.items()):
    print(f"  {s:<35} {n} aziende")
